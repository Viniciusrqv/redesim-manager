"""
cnae_tools.py
-------------
- Classificador de risco a partir da matriz `cnae_risco`.
- Cruzamento com `vigilancia_sanitaria`.
- Extração de CNAEs de PDFs de Cartão CNPJ (pdfplumber + fallback PyMuPDF).
- Regra de consolidação: risco final = maior entre os encontrados
  (Baixo < Médio < Alto). Exige sanitária = SIM se QUALQUER CNAE exigir.
"""
from __future__ import annotations

import re
from typing import Iterable

import pdfplumber

from database import buscar_risco_cnae, buscar_vigilancia


NIVEL_RISCO = {"Baixo": 1, "Médio": 2, "Alto": 3}
RISCO_REVERSO = {v: k for k, v in NIVEL_RISCO.items()}


# ============================================================
# NORMALIZAÇÃO DO CNAE
# ============================================================
# Aceita formatos: 4711-3/02, 47.11-3-02, 4711302, 47113/02
_CNAE_REGEX = re.compile(r"\b\d{2}[\.\-]?\d{2}[\.\-]?\d[/\-]?\d{2}\b")


def normalizar_cnae(raw: str) -> str:
    """Padroniza CNAE para o formato '9999-9/99'."""
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 7:
        return f"{digits[:4]}-{digits[4]}/{digits[5:]}"
    return raw.strip()


# ============================================================
# CLASSIFICAÇÃO
# ============================================================
def classificar_cnae(cnae: str) -> dict:
    """
    Classifica um CNAE buscando:
      1. Match exato da subclasse (7 dígitos, ex '8599-6/04')
      2. Fallback para a classe pai (5 dígitos, ex '85.99-6' — formato da NR-04)

    Retorna {cnae, descricao, risco, grau_risco, exige_sanitaria,
             nivel_sanitaria, risco_sanitario, exige_avcb, grau_avcb,
             ocupacao_it01, area_limite_m2, fonte}.
    Se não encontrado, risco = 'Desconhecido'.
    """
    cnae_norm = normalizar_cnae(cnae)
    cnae_classe = classe_do_cnae(cnae_norm)

    # 1. Tenta match exato da subclasse na matriz de risco
    risco_row = buscar_risco_cnae(cnae_norm)
    # 2. Fallback: classe pai (formato NR-04: '85.99-6')
    if not risco_row:
        risco_row = buscar_risco_cnae(cnae_classe)

    # Mesma lógica pra vigilância sanitária
    vig_row = buscar_vigilancia(cnae_norm)
    if not vig_row:
        vig_row = buscar_vigilancia(cnae_classe)

    # Classificação de Bombeiros (IT-01/CBPMESP)
    # Import local pra evitar ciclo.
    try:
        from database import buscar_bombeiros_cnae as _buscar_bomb
        bomb_row = _buscar_bomb(cnae_norm) or _buscar_bomb(cnae_classe)
    except Exception:
        bomb_row = None

    return {
        "cnae": cnae_norm,
        "cnae_classe": cnae_classe,
        "descricao": (risco_row or {}).get("descricao")
                     or (vig_row or {}).get("descricao")
                     or (bomb_row or {}).get("descricao"),
        "risco": (risco_row or {}).get("risco", "Desconhecido"),
        "grau_risco": (risco_row or {}).get("grau_risco"),
        "fonte": (risco_row or {}).get("fonte"),
        "exige_sanitaria": bool((vig_row or {}).get("exige_licenca", 0)),
        "nivel_sanitaria": (vig_row or {}).get("nivel"),
        "risco_sanitario": (vig_row or {}).get("risco_sanitario"),
        "exige_avcb": bool((bomb_row or {}).get("exige_avcb", 0)),
        "grau_avcb": (bomb_row or {}).get("grau_risco"),
        "ocupacao_it01": (bomb_row or {}).get("ocupacao_it01"),
        "area_limite_m2": (bomb_row or {}).get("area_limite_m2"),
    }


def consolidar(cnaes: Iterable[str]) -> dict:
    """
    Avalia um conjunto de CNAEs e devolve a classificação geral.

    Regras de negócio:
      - Risco final = maior grau entre os CNAEs (Baixo < Médio < Alto).
      - Se QUALQUER CNAE exigir Vigilância Sanitária, o risco vira **Alto**
        automaticamente (a empresa terá que fazer VISA antes da licença).
    """
    detalhes = [classificar_cnae(c) for c in cnaes if c]
    nivel_max = 0
    exige_sanitaria = False
    nivel_sanitaria = None  # Estadual/Municipal/Federal (mais restritivo)
    risco_sanitario_max = 0
    exige_avcb = False
    grau_avcb_max = 0
    ocupacoes_avcb = []

    for d in detalhes:
        nivel_max = max(nivel_max, NIVEL_RISCO.get(d["risco"], 0))
        if d["exige_sanitaria"]:
            exige_sanitaria = True
            if d.get("nivel_sanitaria"):
                nivel_sanitaria = d["nivel_sanitaria"]
            risco_sanitario_max = max(
                risco_sanitario_max,
                NIVEL_RISCO.get(d.get("risco_sanitario"), 0),
            )
        if d.get("exige_avcb"):
            exige_avcb = True
            grau_avcb_max = max(
                grau_avcb_max,
                NIVEL_RISCO.get(d.get("grau_avcb"), 0),
            )
            if d.get("ocupacao_it01"):
                ocupacoes_avcb.append(d["ocupacao_it01"])

    # Regra de negócio: exigir VISA promove pra Alto
    if exige_sanitaria:
        nivel_max = max(nivel_max, NIVEL_RISCO["Alto"])

    risco_final = RISCO_REVERSO.get(nivel_max, "Desconhecido")
    risco_sanitario_consolidado = (
        RISCO_REVERSO.get(risco_sanitario_max) if exige_sanitaria else None
    )
    grau_avcb_consolidado = (
        RISCO_REVERSO.get(grau_avcb_max) if exige_avcb else None
    )

    motivos = []
    if exige_sanitaria:
        motivos.append("CNAE exige Vigilância Sanitária")
    if exige_avcb and grau_avcb_max >= NIVEL_RISCO["Alto"]:
        motivos.append("Atividade de alto risco de incêndio (IT-01)")

    return {
        "risco_final": risco_final,
        "exige_sanitaria": exige_sanitaria,
        "nivel_sanitaria": nivel_sanitaria,
        "risco_sanitario": risco_sanitario_consolidado,
        "exige_avcb": exige_avcb,
        "grau_avcb": grau_avcb_consolidado,
        "ocupacoes_avcb": sorted(set(ocupacoes_avcb)),
        "motivo_alto": " · ".join(motivos) if motivos else None,
        "detalhes": detalhes,
    }


# ============================================================
# EXTRAÇÃO DE PDF (CARTÃO CNPJ)
# ============================================================
def extrair_cnaes_pdf(caminho_pdf: str) -> list[str]:
    """
    Tenta pdfplumber primeiro; se falhar, usa PyMuPDF.
    Retorna lista de CNAEs únicos normalizados.
    """
    texto = ""
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto += (page.extract_text() or "") + "\n"
    except Exception:
        try:
            import fitz  # PyMuPDF
            with fitz.open(caminho_pdf) as doc:
                for page in doc:
                    texto += page.get_text() + "\n"
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"Falha ao extrair PDF: {exc}") from exc

    encontrados = {normalizar_cnae(m) for m in _CNAE_REGEX.findall(texto)}
    encontrados.discard("")
    return sorted(encontrados)


def classe_do_cnae(cnae: str) -> str:
    """
    Converte subclasse CNAE (7 dígitos, ex '8599-6/04') para a classe pai
    (5 dígitos, formato '85.99-6' — como aparece na NR-04).
    Útil pra buscar no banco quando só temos a classe cadastrada.
    """
    if not cnae:
        return ""
    digits = re.sub(r"\D", "", cnae)
    if len(digits) >= 5:
        # '8599604' -> '85.99-6'
        return f"{digits[:2]}.{digits[2:4]}-{digits[4]}"
    return cnae


GRAU_RISCO_PARA_LABEL = {
    1: "Baixo",
    2: "Baixo",
    3: "Médio",
    4: "Alto",
}


def extrair_tabela_nr04_pdf(caminho_pdf: str) -> list[dict]:
    """
    Extrai o Quadro I da NR-04 (Portaria SIT/DSST nº 76/2008 e atualizações):
    Classes CNAE com Grau de Risco de 1 a 4 para dimensionamento do SESMT.

    Formato das linhas: '01.11-3 Cultivo de cereais 3'
    Onde o último número (1-4) é o Grau de Risco.

    Retorna lista de dicts:
        {cnae_classe, descricao, grau_risco (int), risco (Baixo/Médio/Alto), fonte}
    """
    texto_completo = ""
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto_completo += (page.extract_text() or "") + "\n"
    except Exception:
        import fitz
        with fitz.open(caminho_pdf) as doc:
            for page in doc:
                texto_completo += page.get_text() + "\n"

    # Padrão: XX.XX-X  descrição  GR
    # Ex: '01.11-3 Cultivo de cereais 3'
    padrao = re.compile(
        r"^\s*(\d{2}\.\d{2}-\d)\s+(.+?)\s+([1-4])\s*$",
        re.MULTILINE,
    )
    resultados = []
    vistos = set()
    for m in padrao.finditer(texto_completo):
        cnae_classe = m.group(1)
        descricao = re.sub(r"\s+", " ", m.group(2)).strip()
        gr = int(m.group(3))
        if cnae_classe in vistos:
            continue
        vistos.add(cnae_classe)
        resultados.append({
            "cnae_classe": cnae_classe,
            "descricao": descricao,
            "grau_risco": gr,
            "risco": GRAU_RISCO_PARA_LABEL[gr],
            "fonte": "NR-04",
        })
    return resultados


def extrair_tabela_vigilancia_pdf(caminho_pdf: str) -> list[dict]:
    """
    Extrai a tabela CNAE + risco sanitário (ALTO/MÉDIO/BAIXO) de PDFs de
    Portarias/Resoluções de Vigilância Sanitária (ex: Portaria CVS-SP nº 1/2024).

    Para cada CNAE encontrado, retorna:
        {cnae, descricao, risco_sanitario, exige_licenca=1}

    Regras importantes:
    - Blocos marcados como "NÃO COMPREENDE" / "NÃO COMPREENDEM" são
      **exclusões** — CNAEs listados ali não entram no resultado (a menos que
      também apareçam fora do bloco de exclusão como cabeçalho de atividade).
    - Blocos "NÃO SE APLICA" / "EXCETO" recebem o mesmo tratamento.
    """
    texto_completo = ""
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto_completo += (page.extract_text() or "") + "\n"
    except Exception:
        import fitz
        with fitz.open(caminho_pdf) as doc:
            for page in doc:
                texto_completo += page.get_text() + "\n"

    # ------------------------------------------------------------------
    # 1) IDENTIFICAR ZONAS DE EXCLUSÃO  ("NÃO COMPREENDE", "NÃO SE APLICA")
    # ------------------------------------------------------------------
    # Obs.: "EXCETO" não é usado como marcador porque aparece em títulos
    # (ex: "FABRICAÇÃO DE ÓLEOS VEGETAIS EM BRUTO, EXCETO ÓLEO DE MILHO").
    marcadores_excl = re.compile(
        r"N[ÃA]O\s+COMPREENDE[MS]?|N[ÃA]O\s+SE\s+APLICA",
        re.IGNORECASE,
    )
    # Fim da zona: próximo cabeçalho de atividade "CNAE  DESCRIÇÃO_EM_MAIÚSCULAS"
    fim_zona = re.compile(
        r"\n\s*\d{4}-\d/\d{2}\s+[A-ZÀ-Ú][A-ZÀ-Ú\s,\.\-/()&]{8,}",
    )
    zonas_excl: list[tuple[int, int]] = []
    for m in marcadores_excl.finditer(texto_completo):
        inicio = m.end()
        remanescente = texto_completo[inicio:]
        fim_m = fim_zona.search(remanescente)
        fim = inicio + (fim_m.start() if fim_m else min(3000, len(remanescente)))
        zonas_excl.append((inicio, fim))

    def _em_zona_exclusao(pos: int) -> bool:
        return any(ini <= pos < fim for ini, fim in zonas_excl)

    # ------------------------------------------------------------------
    # 2) MAPEAR TODAS AS OCORRÊNCIAS DE CNAE E FILTRAR POR EXCLUSÃO
    # ------------------------------------------------------------------
    ocorrencias: dict[str, list[int]] = {}
    for m in re.finditer(r"\d{4}-\d/\d{2}", texto_completo):
        ocorrencias.setdefault(m.group(0), []).append(m.start())

    cnaes_excluidos: set[str] = set()
    cnaes_validos: set[str] = set()
    for cnae, posicoes in ocorrencias.items():
        fora = [p for p in posicoes if not _em_zona_exclusao(p)]
        if fora:
            cnaes_validos.add(cnae)
        else:
            cnaes_excluidos.add(cnae)

    # ------------------------------------------------------------------
    # 3) EXTRAIR DADOS DOS CNAE VÁLIDOS  (usa só posições fora das zonas)
    # ------------------------------------------------------------------
    encontrados: dict[str, dict] = {}
    blocos = re.split(r"(\d{4}-\d/\d{2})", texto_completo)
    # índice acumulado para saber a posição de cada bloco no texto
    pos_acum = 0
    offsets = [0]
    for b in blocos:
        pos_acum += len(b)
        offsets.append(pos_acum)

    for i in range(1, len(blocos) - 1, 2):
        cnae = blocos[i]
        if cnae in cnaes_excluidos:
            continue  # CNAE só aparece em zona de exclusão
        pos_cnae = offsets[i]
        if _em_zona_exclusao(pos_cnae):
            continue  # Esta ocorrência está em zona de exclusão; pula

        corpo = blocos[i + 1][:1000]

        # 1ª linha após o código = descrição em CAIXA ALTA
        desc = None
        m_desc = re.match(
            r"\s+([A-ZÀ-Ú0-9][A-ZÀ-Ú0-9\s,\-\.;/()&°'’ºª]{5,200})",
            corpo,
        )
        if m_desc:
            desc = re.sub(r"\s+", " ", m_desc.group(1)).strip()

        # Classificação de risco — busca apenas no trecho válido (antes da
        # próxima zona de exclusão ou próximo código)
        fim_busca = len(corpo)
        m_excl = marcadores_excl.search(corpo)
        if m_excl:
            fim_busca = min(fim_busca, m_excl.start())
        trecho_busca = corpo[:min(600, fim_busca)].upper()
        if "ALTO" in trecho_busca:
            risco = "Alto"
        elif re.search(r"M[ÉE]DIO", trecho_busca):
            risco = "Médio"
        elif "BAIXO" in trecho_busca:
            risco = "Baixo"
        else:
            risco = None

        if cnae not in encontrados:
            encontrados[cnae] = {
                "cnae": cnae,
                "descricao": desc,
                "risco_sanitario": risco,
                "exige_licenca": 1,
            }
        elif risco and not encontrados[cnae]["risco_sanitario"]:
            encontrados[cnae]["risco_sanitario"] = risco
        elif desc and not encontrados[cnae]["descricao"]:
            encontrados[cnae]["descricao"] = desc

    # Remove entradas sem risco detectado (provavelmente são referências cruzadas)
    resultado = [v for v in encontrados.values() if v["risco_sanitario"]]
    return sorted(resultado, key=lambda x: x["cnae"])


_LIXO_OCR = re.compile(
    r"^\s*$|^\d{1,2}:\d{2}|P[áa]gina|Brasilia|data e hora|about:blank"
    r"|^[\.\-_,;:]+$|^\d{2}/\d{2}/\d{4}",
    re.IGNORECASE,
)

# Palavras que são cabeçalhos do Cartão CNPJ (nunca são valores reais)
_CABECALHOS_CARTAO = {
    "NUMERO", "NÚMERO", "COMPLEMENTO", "LOGRADOURO", "BAIRRO",
    "DISTRITO", "MUNICIPIO", "MUNICÍPIO", "UF", "CEP", "INSCRICAO",
    "INSCRIÇÃO", "ABERTURA", "DATA", "CADASTRAL", "SITUACAO", "SITUAÇÃO",
    "ESPECIAL", "EMPRESARIAL", "NOME", "RAZAO", "RAZÃO", "SOCIAL",
    "FANTASIA", "TITULO", "TÍTULO", "ESTABELECIMENTO", "NATUREZA",
    "JURIDICA", "JURÍDICA", "CODIGO", "CÓDIGO", "DESCRIGAO", "DESCRIÇÃO",
    "DESCRICAO", "DAATIVIDADE", "ATIVIDADE", "ECONOMICA", "ECONÔMICA",
    "PRINCIPAL", "ATIVIDADES", "SECUNDARIAS", "SECUNDÁRIAS",
    "ELETRONICO", "ELETRÔNICO", "TELEFONE", "ENDERECO", "ENDEREÇO",
    "ENTE", "FEDERATIVO", "RESPONSAVEL", "RESPONSÁVEL", "EFR",
    "MOTIVO", "MATRIZ",
}


def _parece_cabecalho(valor: str) -> bool:
    """True se o valor for uma concatenação de palavras de cabeçalho
    (comum em OCR de tabelas, ex: 'NUMERO COMPLEMENTO')."""
    if not valor:
        return True
    palavras = re.findall(r"[A-ZÀ-Ú]+", valor.upper())
    if not palavras:
        return False
    # Se TODAS as palavras são cabeçalhos, provavelmente é cabeçalho puro
    cabecalhos = sum(1 for p in palavras if p in _CABECALHOS_CARTAO)
    return cabecalhos >= 2 and cabecalhos / len(palavras) >= 0.6


def _extrair_campo(texto: str, *padroes: str) -> str | None:
    """
    Busca o primeiro padrão regex que der match no texto.
    Se o valor estiver vazio ou for cabeçalho, tenta a próxima linha (OCR).
    Rejeita valores que sejam evidentemente não-valores.
    """
    linhas = texto.splitlines()
    for padrao in padroes:
        m = re.search(padrao, texto, re.IGNORECASE)
        if not m:
            continue
        valor = m.group(1).strip()
        valor = re.split(r"\s{2,}", valor)[0].strip()
        # Se valor vazio, lixo, ou cabeçalho, busca próxima linha não-vazia
        if not valor or _LIXO_OCR.search(valor) or _parece_cabecalho(valor):
            pos = m.start()
            linha_idx = texto[:pos].count("\n")
            for prox in linhas[linha_idx + 1:linha_idx + 5]:
                candidato = prox.strip()
                if not candidato or _LIXO_OCR.search(candidato):
                    continue
                if _parece_cabecalho(candidato):
                    continue
                valor = re.split(r"\s{2,}", candidato)[0].strip()
                break
        if valor and not _LIXO_OCR.search(valor) and not _parece_cabecalho(valor):
            return valor
    return None


def _texto_ilegivel(texto: str) -> bool:
    """
    Detecta PDFs cujo texto veio corrompido (sem Unicode mapping).
    Sinais: muitos '(cid:' ou alta densidade de caracteres de controle.
    """
    if not texto or len(texto) < 50:
        return True
    if texto.count("(cid:") > 5:
        return True
    controle = sum(1 for c in texto if ord(c) < 32 and c not in "\n\r\t")
    if controle / max(len(texto), 1) > 0.05:
        return True
    # Pouquíssimas letras reais
    letras = sum(1 for c in texto if c.isalpha())
    if letras / max(len(texto), 1) < 0.20:
        return True
    return False


def _localizar_tesseract_windows() -> str | None:
    """Procura o tesseract.exe nos caminhos padrão de instalação do Windows."""
    import os
    candidatos = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.expanduser(r"~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"),
        os.path.expanduser(r"~\AppData\Local\Tesseract-OCR\tesseract.exe"),
    ]
    for c in candidatos:
        if os.path.isfile(c):
            return c
    return None


def _extrair_texto_com_ocr(caminho_pdf: str) -> tuple[str, str | None]:
    """
    Usa OCR (Tesseract) quando o PDF tem fonte sem Unicode.
    Requer tesseract instalado no sistema (tesseract-ocr no Linux,
    UB-Mannheim/tesseract no Windows).

    Retorna (texto_extraido, idioma_usado). idioma_usado pode ser
    'por', 'eng' ou None (se OCR falhou completamente).
    """
    import os
    import shutil
    import pypdfium2 as pdfium
    import pytesseract

    # Se o tesseract não está no PATH, tenta localizar nos caminhos padrão do Windows
    if os.name == "nt" and not shutil.which("tesseract"):
        caminho = _localizar_tesseract_windows()
        if caminho:
            pytesseract.pytesseract.tesseract_cmd = caminho

    texto = ""
    idioma_usado = None
    pdf = pdfium.PdfDocument(caminho_pdf)
    try:
        # tenta português, cai para inglês se por.traineddata não estiver
        for lang in ("por", "eng"):
            try:
                buffer = ""
                for page in pdf:
                    # scale=4 melhora o reconhecimento do CNPJ no cabeçalho
                    img = page.render(scale=4).to_pil()
                    # PSM 6 ("uniform block of text") preserva o layout em
                    # linhas tabulares do Cartão CNPJ — é o que melhor
                    # recupera o CNPJ e a linha endereço+UF.
                    # PSM 11 ("sparse text") é usado como complemento para
                    # cabeçalhos isolados que o PSM 6 não alcança.
                    txt6 = pytesseract.image_to_string(
                        img, lang=lang, config="--psm 6"
                    )
                    txt11 = pytesseract.image_to_string(
                        img, lang=lang, config="--psm 11"
                    )
                    buffer += txt6 + "\n[--PSM11--]\n" + txt11 + "\n"
                    img.close()
                if buffer.strip():
                    texto = buffer
                    idioma_usado = lang
                    break
            except pytesseract.TesseractError:
                continue
    finally:
        pdf.close()
    return texto, idioma_usado


def _parsear_endereco_cartao(texto: str) -> dict:
    """
    Parser específico para o bloco de endereço do Cartão CNPJ.

    Layout típico (em 3 colunas):
        LOGRADOURO            NUMERO        COMPLEMENTO
        AL RIO NEGRO          123           SALA 2020

        CEP            BAIRRO/DISTRITO      MUNICIPIO       UF
        06454-000      ALPHAVILLE           BARUERI         SP

    OCR frequentemente embaralha tudo em linhas diferentes.
    A estratégia: localizar a linha de cabeçalho e parsear a próxima linha
    usando heurísticas (números = numero/CEP, UF = 2 letras maiúsculas).
    """
    linhas = [l.strip() for l in texto.splitlines()]

    # Índices dos cabeçalhos
    idx_logradouro = None
    idx_cep = None
    for i, linha in enumerate(linhas):
        up = linha.upper()
        if idx_logradouro is None and "LOGRADOURO" in up:
            idx_logradouro = i
        if idx_cep is None and "CEP" in up and ("BAIRRO" in up or "MUNICIPIO" in up or "MUNICÍPIO" in up):
            idx_cep = i

    resultado = {
        "logradouro": None, "numero": None, "complemento": None,
        "cep": None, "bairro": None, "municipio": None, "uf": None,
    }

    # --- Linha LOGRADOURO ---
    if idx_logradouro is not None:
        for offset in range(1, 5):
            if idx_logradouro + offset >= len(linhas):
                break
            valor = linhas[idx_logradouro + offset]
            if not valor or _LIXO_OCR.search(valor) or _parece_cabecalho(valor):
                continue
            # O OCR pode dar: "AL RIO NEGRO 123 SALA 2020"
            # ou: "AL RIO NEGRO SALA 2020" (sem número, só complemento)
            partes = valor.split()
            # Palavras que indicam que o número seguinte é complemento,
            # não número da rua
            TOKENS_COMPL = {
                "SALA", "SALAS", "BLOCO", "BL", "APTO", "AP",
                "ANDAR", "CONJ", "CONJUNTO", "LOJA", "LT", "QD",
                "CASA", "FUNDOS", "EDIF", "EDIFICIO", "EDIFÍCIO",
                "TORRE", "PAV", "PAVIMENTO",
            }
            num_idx = None
            for j, p in enumerate(partes):
                if re.fullmatch(r"\d{1,6}", p):
                    # se palavra anterior é indicador de complemento,
                    # este número NÃO é número da rua
                    if j > 0 and partes[j - 1].upper().rstrip(".") in TOKENS_COMPL:
                        continue
                    num_idx = j
                    break
            if num_idx is not None:
                resultado["logradouro"] = " ".join(partes[:num_idx]).strip() or None
                resultado["numero"] = partes[num_idx]
                if num_idx + 1 < len(partes):
                    resultado["complemento"] = " ".join(partes[num_idx + 1:])
            else:
                # Sem número da rua — procura onde começa o complemento
                idx_compl = None
                for j, p in enumerate(partes):
                    if p.upper().rstrip(".") in TOKENS_COMPL:
                        idx_compl = j
                        break
                if idx_compl is not None:
                    resultado["logradouro"] = " ".join(partes[:idx_compl]).strip() or None
                    resultado["complemento"] = " ".join(partes[idx_compl:])
                else:
                    resultado["logradouro"] = valor
            break

    # --- Linha CEP ---
    if idx_cep is not None:
        for offset in range(1, 5):
            if idx_cep + offset >= len(linhas):
                break
            valor = linhas[idx_cep + offset]
            if not valor or _LIXO_OCR.search(valor) or _parece_cabecalho(valor):
                continue
            # Procura CEP (00000-000 ou 00.000-000)
            m_cep = re.search(r"(\d{2}\.?\d{3}\-?\d{3})", valor)
            if m_cep:
                resultado["cep"] = re.sub(r"\D", "", m_cep.group(1))
                if len(resultado["cep"]) == 8:
                    resultado["cep"] = f"{resultado['cep'][:5]}-{resultado['cep'][5:]}"
                resto = valor[m_cep.end():].strip()
            else:
                resto = valor
            # Procura UF no final (duas letras maiúsculas isoladas)
            m_uf = re.search(r"\b([A-Z]{2})\s*$", resto)
            if m_uf and m_uf.group(1) not in ("DE", "DO", "DA"):
                resultado["uf"] = m_uf.group(1)
                resto = resto[:m_uf.start()].strip()
            # Resto = bairro + município — heurística: a última palavra
            # é o município (ex: "ALPHAVILLE CENTRO BARUERI" -> bairro =
            # "ALPHAVILLE CENTRO", município = "BARUERI"). Cidades compostas
            # (ex: SAO PAULO) requerem lista de exceções, mas priorizamos
            # o caso mais comum.
            partes = resto.split()
            if len(partes) >= 2:
                resultado["bairro"] = " ".join(partes[:-1])
                resultado["municipio"] = partes[-1]
            elif len(partes) == 1:
                resultado["municipio"] = partes[0]
            break

    return resultado


def extrair_dados_cartao_cnpj(caminho_pdf: str) -> dict:
    """
    Extrai todos os dados relevantes de um Cartão CNPJ (Receita Federal):
    CNPJ, Razão Social, Nome Fantasia, Endereço completo, Município, UF,
    CEP, CNAE principal e secundários, Situação Cadastral, Data de abertura.

    Fluxo: pdfplumber → PyMuPDF → OCR (Tesseract).
    """
    texto = ""
    usou_ocr = False
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto += (page.extract_text() or "") + "\n"
    except Exception:
        try:
            import fitz
            with fitz.open(caminho_pdf) as doc:
                for page in doc:
                    texto += page.get_text() + "\n"
        except Exception:
            texto = ""

    # PyMuPDF (fitz) tambem quando o texto veio ilegivel (nao so em excecao):
    # fonte sem Unicode que o pdfplumber le como lixo as vezes sai certa no fitz.
    if _texto_ilegivel(texto):
        try:
            import fitz
            _t_fitz = ""
            with fitz.open(caminho_pdf) as _doc:
                for _pg in _doc:
                    _t_fitz += _pg.get_text() + "\n"
            if not _texto_ilegivel(_t_fitz):
                texto = _t_fitz
        except Exception:
            pass

    # Fallback para OCR se o texto estiver ilegível
    idioma_ocr = None
    if _texto_ilegivel(texto):
        try:
            texto_ocr, idioma_ocr = _extrair_texto_com_ocr(caminho_pdf)
            if texto_ocr.strip():
                texto = texto_ocr
                usou_ocr = True
        except Exception as exc:
            # Mantém o texto original e sinaliza falha de OCR
            texto = (texto or "") + f"\n[OCR_ERRO:{exc}]"

    # Normaliza espaços/linhas mas mantém quebra de linha
    texto_limpo = re.sub(r"[ \t]+", " ", texto)

    # ---------- CNPJ ----------
    cnpj = None
    # Formato ideal: 00.000.000/0000-00
    cnpj_match = re.search(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", texto_limpo)
    if cnpj_match:
        cnpj = cnpj_match.group()
    else:
        # OCR costuma distorcer os separadores (vírgula no lugar de ponto,
        # espaço extra, . no lugar de /). Aceita qualquer não-dígito entre
        # os grupos, desde que dê 14 dígitos no total.
        m = re.search(
            r"\b(\d{2})[\.\,\-\s]{0,2}(\d{3})[\.\,\-\s]{0,2}(\d{3})"
            r"[\.\,\-\s/]{0,3}(\d{4})[\.\,\-\s]{0,2}(\d{2})\b",
            texto_limpo,
        )
        if m:
            cnpj = f"{m.group(1)}.{m.group(2)}.{m.group(3)}/{m.group(4)}-{m.group(5)}"

    # ---------- Razão Social ----------
    razao = _extrair_campo(
        texto_limpo,
        r"NOME EMPRESARIAL\s*[:\-]?\s*([^\n]+)",
        r"RAZ[ÃA]O SOCIAL\s*[:\-]?\s*([^\n]+)",
    )

    # ---------- Nome Fantasia ----------
    fantasia = _extrair_campo(
        texto_limpo,
        r"T[ÍI]TULO DO ESTABELECIMENTO[^:]*:\s*([^\n]+)",
        r"NOME FANTASIA\s*[:\-]?\s*([^\n]+)",
        # Layout PSM 6 junta o PORTE na mesma linha do fantasia.
        # Pegamos a linha inteira e tiramos o PORTE depois.
        r"T[ÍI]TULO\s+DO\s+ESTABELECIMENTO[^\n]*\n+([^\n]+)",
    )
    if fantasia and fantasia.startswith("("):
        fantasia = None
    # Remove o sufixo de PORTE (ME / EPP / DEMAIS) que às vezes fica
    # grudado no nome fantasia por causa do layout em colunas.
    if fantasia:
        fantasia = re.sub(
            r"\s+(ME|EPP|DEMAIS)\s*$", "", fantasia.strip(), flags=re.IGNORECASE
        ).strip() or None

    # ---------- Endereço ----------
    # Quando o texto vem de OCR, o layout em 3 colunas quebra o regex simples.
    # Nesse caso usamos o parser posicional diretamente.
    if usou_ocr:
        end = _parsear_endereco_cartao(texto_limpo)
        logradouro = end["logradouro"]
        numero = end["numero"]
        complemento = end["complemento"]
        bairro = end["bairro"]
        cep = end["cep"]
        municipio = end["municipio"]
        uf = end["uf"]
    else:
        logradouro = _extrair_campo(texto_limpo, r"LOGRADOURO\s*[:\-]?\s*([^\n]+)")
        numero = _extrair_campo(texto_limpo, r"N[ÚU]MERO\s*[:\-]?\s*([^\n]+)")
        complemento = _extrair_campo(texto_limpo, r"COMPLEMENTO\s*[:\-]?\s*([^\n]+)")
        bairro = _extrair_campo(texto_limpo, r"BAIRRO[^:\n]*[:\-]?\s*([^\n]+)")
        cep = _extrair_campo(texto_limpo, r"CEP\s*[:\-]?\s*([\d\.\-]+)")
        municipio = _extrair_campo(texto_limpo, r"MUNIC[ÍI]PIO\s*[:\-]?\s*([^\n]+)")
        uf = _extrair_campo(texto_limpo, r"\bUF\s*[:\-]?\s*([A-Z]{2})\b")

        # Se mesmo em texto nativo a extração falhou, tenta parser posicional
        if not any([logradouro, numero, bairro, cep, municipio, uf]):
            end = _parsear_endereco_cartao(texto_limpo)
            logradouro = end["logradouro"]
            numero = end["numero"]
            complemento = end["complemento"]
            bairro = end["bairro"]
            cep = end["cep"]
            municipio = end["municipio"]
            uf = end["uf"]

    # Fallback de UF: busca sigla válida APÓS o município (mais confiável),
    # ou no contexto da linha do CEP. Evita confusão com abreviações de
    # logradouro como "AL" (Alameda), "AV" (Avenida).
    if not uf:
        UFS_VALIDAS = {
            "AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS",
            "MG","PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC",
            "SP","SE","TO",
        }
        candidatos = []
        # 1. Procura UF logo após o município (ex: "BARUERI SP")
        if municipio:
            mun_escaped = re.escape(municipio)
            m = re.search(
                rf"{mun_escaped}\s+([A-Z]{{2}})\b",
                texto_limpo, re.IGNORECASE,
            )
            if m and m.group(1).upper() in UFS_VALIDAS:
                uf = m.group(1).upper()
        # 2. Procura UF na linha que contém o CEP (busca mais abrangente)
        if not uf and cep:
            cep_digits = re.sub(r"\D", "", cep)
            # aceita CEP com ou sem traço e com ou sem ponto
            cep_pat = rf"{cep_digits[:2]}\.?{cep_digits[2:5]}\-?{cep_digits[5:]}"
            m_cep = re.search(cep_pat, texto_limpo)
            if m_cep:
                # pega os próximos 200 chars após o CEP
                trecho = texto_limpo[m_cep.end():m_cep.end() + 200]
                # para na primeira quebra de linha dupla (fim da seção)
                trecho = trecho.split("\n\n")[0]
                for m in re.finditer(r"(?<![A-Z])([A-Z]{2})(?![A-Z])", trecho):
                    if m.group(1) in UFS_VALIDAS:
                        uf = m.group(1)
                        break

    # Monta endereço único para exibição
    partes = [p for p in [logradouro, numero, complemento, bairro] if p]
    endereco = ", ".join(partes) if partes else None
    if cep and endereco:
        endereco = f"{endereco} - CEP {cep}"

    # ---------- Situação / Abertura / Porte ----------
    # OCR frequentemente troca C por G (SITUAGAO ao invés de SITUACAO).
    # A situação (ATIVA/SUSPENSA/...) costuma vir na linha IMEDIATAMENTE
    # abaixo do cabeçalho "SITUA[CG]AO CADASTRAL". Procuramos a primeira
    # palavra válida numa janela de até 3 linhas após o cabeçalho.
    situacao = None
    m_sit = re.search(
        r"SITUA[ÇCG][ÃA]O\s+CADASTRAL[^\n]*\n((?:[^\n]*\n){0,3})",
        texto_limpo,
        re.IGNORECASE,
    )
    if m_sit:
        bloco = m_sit.group(1)
        valida = {"ATIVA", "SUSPENSA", "INAPTA", "BAIXADA", "NULA"}
        for palavra in re.findall(r"[A-ZÀ-Úa-zà-ú]+", bloco.upper()):
            if palavra in valida:
                situacao = palavra
                break

    # DATA DE ABERTURA: exige formato dd/mm/aaaa. Procura na mesma linha
    # do cabeçalho (quando vem inline) ou na linha imediatamente abaixo.
    # O `[^\n\d]*` impede que o regex "pule" linhas intermediárias e
    # acabe pegando a data de emissão do cartão lá no rodapé.
    abertura = None
    m_ab = re.search(
        r"DATA\s+DE\s+ABERTURA[^\n\d]*(?:\n[^\n]*?)?(\d{2}/\d{2}/\d{4})",
        texto_limpo,
        re.IGNORECASE,
    )
    if m_ab:
        abertura = m_ab.group(1)
    else:
        # Fallback: pega a data da SITUAÇÃO CADASTRAL (comum quando a
        # DATA DE ABERTURA está ilegível no OCR).
        m_sit_date = re.search(
            r"SITUA[ÇCG][ÃA]O\s+CADASTRAL[^\n\d]*\n[^\n]*?(\d{2}/\d{2}/\d{4})",
            texto_limpo,
            re.IGNORECASE,
        )
        if m_sit_date:
            abertura = m_sit_date.group(1)

    # PORTE: se presente, vem como "ME", "EPP" ou "DEMAIS". Filtra pra
    # não pegar o nome fantasia (que pode estar na mesma coluna).
    porte = _extrair_campo(texto_limpo, r"\bPORTE\s*[:\-]?\s*([^\n]+)")
    if porte:
        m_p = re.search(r"\b(ME|EPP|DEMAIS)\b", porte.upper())
        porte = m_p.group(1) if m_p else None

    # ---------- CNPJ (fallback se OCR corrompeu) ----------
    if cnpj is None and usou_ocr:
        # Procura 14 dígitos consecutivos em qualquer ordem
        m = re.search(r"(\d{2})\D*(\d{3})\D*(\d{3})\D*(\d{4})\D*(\d{2})", texto_limpo)
        if m:
            cnpj = f"{m.group(1)}.{m.group(2)}.{m.group(3)}/{m.group(4)}-{m.group(5)}"

    # ---------- CNAEs ----------
    # CNAE principal costuma vir logo após "ATIVIDADE ECONÔMICA PRINCIPAL"
    cnae_principal = None
    m = re.search(
        r"ATIVIDADE ECON[ÔO]MICA PRINCIPAL[^0-9]*?(\d{2}\.?\d{2}-?\d-?\d{2})",
        texto_limpo, re.IGNORECASE,
    )
    if m:
        cnae_principal = normalizar_cnae(m.group(1))

    todos_cnaes = sorted({normalizar_cnae(c) for c in _CNAE_REGEX.findall(texto_limpo)})
    todos_cnaes = [c for c in todos_cnaes if c]

    # Reordena para que o principal venha primeiro
    if cnae_principal and cnae_principal in todos_cnaes:
        todos_cnaes.remove(cnae_principal)
        todos_cnaes = [cnae_principal] + todos_cnaes

    return {
        "cnpj": cnpj,
        "razao_social": razao,
        "nome_fantasia": fantasia,
        "logradouro": logradouro,
        "numero": numero,
        "complemento": complemento,
        "bairro": bairro,
        "cep": cep,
        "municipio": municipio,
        "uf": uf,
        "endereco": endereco,
        "situacao": situacao,
        "data_abertura": abertura,
        "porte": porte,
        "cnae_principal": cnae_principal,
        "cnaes": todos_cnaes,
        "texto": texto,
        "usou_ocr": usou_ocr,
        "idioma_ocr": idioma_ocr,
    }


# ============================================================
# EXTRAÇÃO DE AVCB / CLCB (CBPMESP)
# ============================================================
def extrair_dados_avcb(caminho_pdf: str) -> dict:
    """
    Extrai os campos relevantes de um AVCB (Auto de Vistoria) ou CLCB
    (Certificado de Licença) emitido pelo Corpo de Bombeiros de SP.

    Retorna um dict com:
        tipo              "AVCB" | "CLCB" | None
        numero            número do documento (string)
        data_emissao      dd/mm/aaaa
        data_vencimento   dd/mm/aaaa
        razao_social      string
        cnpj              00.000.000/0000-00
        endereco          string composta (logradouro + número + bairro + município/UF)
        ocupacao          letra (A..M) — grupo IT-01
        divisao           string (ex "D-3", "F-6")
        descricao_ocupacao descrição textual da ocupação
        area_construida   float (m²)
        altura            float (m), quando detectável
        texto             texto bruto (debug)
        usou_ocr          bool

    Estratégia: pdfplumber → fallback OCR (mesma função do Cartão CNPJ).
    AVCB/CLCB do CBPMESP costumam ter layout tabular bem consistente,
    então procuramos os cabeçalhos clássicos.
    """
    texto = ""
    usou_ocr = False
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto += (page.extract_text() or "") + "\n"
    except Exception:
        try:
            import fitz
            with fitz.open(caminho_pdf) as doc:
                for page in doc:
                    texto += page.get_text() + "\n"
        except Exception:
            texto = ""

    idioma_ocr = None
    if _texto_ilegivel(texto):
        try:
            texto_ocr, idioma_ocr = _extrair_texto_com_ocr(caminho_pdf)
            if texto_ocr.strip():
                texto = texto_ocr
                usou_ocr = True
        except Exception as exc:
            texto = (texto or "") + f"\n[OCR_ERRO:{exc}]"

    texto_limpo = re.sub(r"[ \t]+", " ", texto)
    up = texto_limpo.upper()

    # ---------- TIPO ----------
    tipo = None
    if re.search(r"AUTO\s+DE\s+VISTORIA", up) or re.search(r"\bAVCB\b", up):
        tipo = "AVCB"
    elif re.search(r"CERTIFICADO\s+DE\s+LICEN", up) or re.search(r"\bCLCB\b", up):
        tipo = "CLCB"

    # ---------- NÚMERO DO DOCUMENTO ----------
    # Possíveis rótulos: "Nº AVCB", "Nº do Auto", "Nº do Certificado",
    # "Número do AVCB", "AVCB Nº", "CLCB Nº", "Nº do processo", e o
    # pior caso: "No do AVCB" (OCR lê "º" como "o").
    # Regex evita pegar "NUMERO: 100" do bloco de endereço.
    numero = None
    # prefixo que aceita Nº, N°, No, NO, Nro, Numero/Número etc.
    # (texto já está em UPPER, então precisa aceitar 'O' maiúsculo)
    N = r"(?:N[ºOo°ª\.\s]{0,3}|N[ÚU]MERO|NRO|NUM\.?)"
    for pat in (
        rf"{N}\s*(?:DO\s+)?AVCB\s*[:\-]?\s*([\w\./-]+)",
        rf"{N}\s*(?:DO\s+)?CLCB\s*[:\-]?\s*([\w\./-]+)",
        rf"AVCB\s*{N}\s*[:\-]?\s*([\w\./-]+)",
        rf"CLCB\s*{N}\s*[:\-]?\s*([\w\./-]+)",
        rf"{N}\s*(?:DO\s+)?AUTO(?:\s+DE\s+VISTORIA)?\s*[:\-]?\s*([\w\./-]+)",
        rf"{N}\s*(?:DO\s+)?CERTIFICADO(?:\s+DE\s+LICEN[ÇC]A)?\s*[:\-]?\s*([\w\./-]+)",
    ):
        m = re.search(pat, up)
        if m:
            cand = m.group(1).strip(".-/")
            # precisa ter ao menos 3 dígitos pra ser número de AVCB válido
            if re.search(r"\d{3,}", cand):
                numero = cand
                break

    # ---------- DATAS ----------
    # Busca todas as datas dd/mm/aaaa e tenta rotular
    # pelas palavras próximas (EMISSÃO / VALIDADE / VENCIMENTO).
    data_emissao = None
    data_vencimento = None

    # rótulos conhecidos (em UPPER)
    for rot in ("VALIDADE", "VENCIMENTO", "VALIDO ATE", "V[ÁA]LIDO\\s+AT[ÉE]"):
        m = re.search(
            rf"{rot}[^\n\d]*(\d{{2}}/\d{{2}}/\d{{4}})",
            up,
        )
        if m:
            data_vencimento = m.group(1)
            break

    for rot in ("EMISS[ÃA]O", "EMITIDO", "DATA\\s+DA\\s+EMISS"):
        m = re.search(
            rf"{rot}[^\n\d]*(\d{{2}}/\d{{2}}/\d{{4}})",
            up,
        )
        if m:
            data_emissao = m.group(1)
            break

    # Fallback: se achou só uma data e é maior que hoje, é vencimento
    if not data_emissao or not data_vencimento:
        todas_datas = re.findall(r"\b(\d{2}/\d{2}/\d{4})\b", texto_limpo)
        if todas_datas:
            if not data_emissao:
                data_emissao = todas_datas[0]
            if not data_vencimento and len(todas_datas) > 1:
                data_vencimento = todas_datas[-1]

    # ---------- RAZÃO SOCIAL ----------
    razao = _extrair_campo(
        texto_limpo,
        r"RAZ[ÃA]O\s+SOCIAL\s*[:\-]?\s*([^\n]+)",
        r"PROPRIET[ÁA]RIO(?:\(A\))?\s*[:\-]?\s*([^\n]+)",
        r"RESPONS[ÁA]VEL\s*[:\-]?\s*([^\n]+)",
    )

    # ---------- CNPJ ----------
    cnpj = None
    m = re.search(r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b", texto_limpo)
    if m:
        cnpj = m.group()
    else:
        m = re.search(
            r"(\d{2})[\.\,\-\s]{0,2}(\d{3})[\.\,\-\s]{0,2}(\d{3})"
            r"[\.\,\-\s/]{0,3}(\d{4})[\.\,\-\s]{0,2}(\d{2})",
            texto_limpo,
        )
        if m:
            cnpj = f"{m.group(1)}.{m.group(2)}.{m.group(3)}/{m.group(4)}-{m.group(5)}"

    # ---------- ENDEREÇO ----------
    logradouro = _extrair_campo(texto_limpo, r"LOGRADOURO\s*[:\-]?\s*([^\n]+)",
                                r"ENDERE[ÇC]O\s*[:\-]?\s*([^\n]+)")
    numero_end = _extrair_campo(texto_limpo, r"N[ÚU]MERO\s*[:\-]?\s*([^\n]+)")
    bairro = _extrair_campo(texto_limpo, r"BAIRRO[^:\n]*[:\-]?\s*([^\n]+)")
    municipio = _extrair_campo(texto_limpo, r"MUNIC[ÍI]PIO\s*[:\-]?\s*([^\n]+)")
    uf = _extrair_campo(texto_limpo, r"\bUF\s*[:\-]?\s*([A-Z]{2})\b")
    cep = _extrair_campo(texto_limpo, r"CEP\s*[:\-]?\s*([\d\.\-]+)")
    partes = [p for p in [logradouro, numero_end, bairro] if p]
    endereco = ", ".join(partes) if partes else None
    if municipio and uf:
        endereco = (endereco + " — " if endereco else "") + f"{municipio}/{uf}"
    elif municipio:
        endereco = (endereco + " — " if endereco else "") + municipio
    if cep:
        endereco = (endereco or "") + f" - CEP {cep}"
    endereco = endereco.strip(" -—,") if endereco else None

    # ---------- OCUPAÇÃO IT-01 ----------
    # O IT-01 classifica em grupos A..M e divisões (ex: D-3, F-6, A-2).
    # Formatos comuns em AVCB: "Ocupação: D-3 Serviços profissionais"
    # ou "Divisão D-3 (Serviços profissionais)"
    ocupacao = None
    divisao = None
    descricao_ocupacao = None

    m = re.search(
        r"(?:OCUPA[ÇC][ÃA]O|DIVIS[ÃA]O|GRUPO)[^\n]*?"
        r"([A-M])\s*[-–]\s*(\d{1,2})"
        r"(?:[^\n]*?[-–:]\s*([^\n]+))?",
        up,
    )
    if m:
        ocupacao = m.group(1)
        divisao = f"{m.group(1)}-{m.group(2)}"
        if m.group(3):
            # descrição em texto original (preserva case)
            start = m.start(3)
            end = m.end(3)
            descricao_ocupacao = texto_limpo[start:end].strip(" :-—,.")

    # Fallback: padrão solto tipo "D-3" isolado
    if not divisao:
        m = re.search(r"\b([A-M])\s*[-–]\s*(\d{1,2})\b", up)
        if m:
            ocupacao = m.group(1)
            divisao = f"{m.group(1)}-{m.group(2)}"

    # ---------- ÁREA CONSTRUÍDA ----------
    area = None
    # Aceita "Área construída: 350,00 m²" / "Área Total: 1.234,56 m2"
    m = re.search(
        r"[ÁA]REA\s+(?:CONSTRU[ÍI]DA|TOTAL|EDIFIC\w+)[^\n\d]*"
        r"([\d\.\,]+)\s*m",
        up,
        re.IGNORECASE,
    )
    if m:
        area_str = m.group(1).replace(".", "").replace(",", ".")
        try:
            area = float(area_str)
        except ValueError:
            area = None

    # ---------- ALTURA ----------
    altura = None
    m = re.search(
        r"ALTURA[^\n\d]*([\d\.\,]+)\s*m(?!²|2)",
        up,
        re.IGNORECASE,
    )
    if m:
        try:
            altura = float(m.group(1).replace(".", "").replace(",", "."))
        except ValueError:
            altura = None

    return {
        "tipo": tipo,
        "numero": numero,
        "data_emissao": data_emissao,
        "data_vencimento": data_vencimento,
        "razao_social": razao,
        "cnpj": cnpj,
        "endereco": endereco,
        "municipio": municipio,
        "uf": uf,
        "cep": cep,
        "ocupacao": ocupacao,
        "divisao": divisao,
        "descricao_ocupacao": descricao_ocupacao,
        "area_construida": area,
        "altura": altura,
        "texto": texto,
        "usou_ocr": usou_ocr,
        "idioma_ocr": idioma_ocr,
    }


# =====================================================
# EXTRATOR GENÉRICO DE DOCUMENTOS COM VENCIMENTO
# =====================================================
# Suporta:
#   - CND Federal (RFB / PGFN)
#   - CND FGTS (Certificado de Regularidade — Caixa)
#   - CNDT (Certidão Negativa de Débitos Trabalhistas — TST)
#   - CND Estadual / CND Municipal (heurística)
#   - Alvará de Funcionamento (municipal)
#   - Alvará Sanitário / Licença Sanitária
#   - Licença Ambiental
#   - Outros (fallback — tenta achar a data de validade mais provável)
#
# Estratégia: pdfplumber → fitz (PyMuPDF) → OCR. Depois aplica
# detectores de tipo por keywords no texto uppercase e extrai
# os campos usando regex específico para cada tipo + um fallback
# genérico pra data de vencimento.

def _extrair_texto_pdf_ou_ocr(caminho_pdf: str) -> tuple[str, bool, str | None]:
    """Pipeline padrão: pdfplumber → fitz → OCR. Retorna (texto, usou_ocr, idioma_ocr)."""
    texto = ""
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                texto += (page.extract_text() or "") + "\n"
    except Exception:
        pass
    if not texto.strip() or _texto_ilegivel(texto):
        try:
            import fitz
            with fitz.open(caminho_pdf) as doc:
                buffer = ""
                for page in doc:
                    buffer += page.get_text() + "\n"
            if buffer.strip() and not _texto_ilegivel(buffer):
                texto = buffer
        except Exception:
            pass

    usou_ocr = False
    idioma_ocr = None
    if _texto_ilegivel(texto):
        try:
            texto_ocr, idioma_ocr = _extrair_texto_com_ocr(caminho_pdf)
            if texto_ocr.strip():
                texto = texto_ocr
                usou_ocr = True
        except Exception as exc:
            texto = (texto or "") + f"\n[OCR_ERRO:{exc}]"
    return texto, usou_ocr, idioma_ocr


def _detectar_tipo_documento(up: str) -> str:
    """Detecta o tipo do documento analisando palavras-chave no texto (já em UPPER)."""
    # Ordem importa — do mais específico pro mais genérico
    # --- Bombeiros (AVCB/CLCB) — muito específico, vai primeiro ---
    if re.search(r"AUTO\s+DE\s+VISTORIA\s+DO\s+CORPO\s+DE\s+BOMBEIROS", up) \
            or re.search(r"\bAVCB\b", up):
        return "AVCB"
    if re.search(r"CERTIFICADO\s+DE\s+LICEN[ÇC]A\s+DO\s+CORPO\s+DE\s+BOMBEIROS", up) \
            or re.search(r"\bCLCB\b", up):
        return "CLCB"
    if ("FGTS" in up and ("REGULARIDADE" in up or "CRF" in up)) \
            or "CERTIFICADO DE REGULARIDADE DO FGTS" in up:
        return "CND FGTS"
    if "CNDT" in up or "DÉBITOS TRABALHISTAS" in up \
            or "DEBITOS TRABALHISTAS" in up \
            or ("TRIBUNAL SUPERIOR DO TRABALHO" in up and "CERTIDÃO" in up):
        return "CNDT (Trabalhista)"
    if "PROCURADORIA-GERAL DA FAZENDA NACIONAL" in up \
            or ("RECEITA FEDERAL" in up and "CERTIDÃO" in up) \
            or ("PGFN" in up and "CERTIDÃO" in up) \
            or "TRIBUTOS FEDERAIS" in up:
        return "CND Federal"
    # Alvará Sanitário / Licença Sanitária (inclui variações de OCR sem acento)
    _markers_sanit = (
        "LICENÇA SANITÁRIA", "LICENCA SANITARIA",
        "ALVARÁ SANITÁRIO", "ALVARA SANITARIO",
        "VIGILÂNCIA SANITÁRIA", "VIGILANCIA SANITARIA",
        "CEVS",  # número CEVS só aparece em documentos de Vigilância Sanitária
    )
    if any(mk in up for mk in _markers_sanit):
        return "Alvará Sanitário"
    if "LICENÇA AMBIENTAL" in up or "LICENCA AMBIENTAL" in up \
            or "CETESB" in up:
        return "Licença Ambiental"
    if "ALVARÁ DE FUNCIONAMENTO" in up or "ALVARA DE FUNCIONAMENTO" in up \
            or "AUTO DE LICENÇA DE FUNCIONAMENTO" in up \
            or "LICENÇA DE FUNCIONAMENTO" in up:
        return "Alvará de Funcionamento"
    if "SECRETARIA DA FAZENDA" in up and "CERTIDÃO" in up:
        return "CND Estadual"
    if ("PREFEITURA" in up or "MUNICÍPIO" in up or "MUNICIPAL" in up) \
            and "CERTIDÃO" in up:
        return "CND Municipal"
    if "CERTIFICADO DIGITAL" in up or "ICP-BRASIL" in up:
        return "Certificado Digital"
    if "CONTRATO SOCIAL" in up:
        return "Contrato Social"
    return "Outro"


_DATA_RE = r"(\d{2}/\d{2}/\d{4})"


def _extrair_cnpj(texto: str) -> str | None:
    # formato com máscara
    m = re.search(r"(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", texto)
    if m:
        return m.group(1)
    # 14 dígitos contínuos — aplica máscara
    m = re.search(r"(?<!\d)(\d{14})(?!\d)", texto)
    if m:
        d = m.group(1)
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return None


def _extrair_razao_social(texto: str, cnpj: str | None) -> str | None:
    """
    Tenta pegar a razão social a partir de rótulos, da vizinhança do CNPJ,
    e (em docs como Alvará Sanitário) reconstruindo linhas adjacentes que
    o OCR cortou ao ler um layout tabular de 2 colunas.
    """
    def _limpar(rs: str) -> str:
        rs = rs.strip(" :\t—-|*")
        # remove rótulos que vêm colados no começo
        rs = re.sub(
            r"^\s*(Raz[ãa]o\s+Social|Nome\s+Empresarial|Nome|Contribuinte)\s*[:\s-]+\s*",
            "", rs, flags=re.IGNORECASE,
        )
        # remove junk de fim de coluna (vinha de layout tabular)
        rs = re.sub(r"\s*CNPJ.*$", "", rs, flags=re.IGNORECASE)
        rs = re.sub(r"\s*ALBERGANTE.*$", "", rs, flags=re.IGNORECASE)
        rs = re.sub(r"\s*(NOME\s+FANTASIA|DETALHE|PAGINA|P[ÁA]GINA)[:\s].*$",
                    "", rs, flags=re.IGNORECASE)
        rs = re.sub(r"\s+", " ", rs).strip(" :\t—-|*")
        return rs

    candidatos_globais: list[str] = []

    # 1) Rótulos explícitos — mesma linha
    for padrao in [
        r"(?:Raz[ãa]o\s+Social|Nome\s+Empresarial|Nome)[:\s]+([^\n]{4,120})",
        r"Contribuinte[:\s]+([^\n]{4,120})",
    ]:
        m = re.search(padrao, texto, re.IGNORECASE)
        if m:
            rs = _limpar(m.group(1))
            if 3 < len(rs) < 130 and sum(c.isalpha() for c in rs) > 4:
                candidatos_globais.append(rs)

    # 2) Nome empresarial all-caps terminando em LTDA/EIRELI/ME/EPP/S.A./MEI
    #    Achatamos o texto (newlines → espaço) e removemos barras de coluna
    #    pra o regex atravessar múltiplas linhas.
    texto_flat = re.sub(r"[|*•·—\-]", " ", texto)
    texto_flat = re.sub(r"\s+", " ", texto_flat)
    up_flat = texto_flat.upper()

    ruins = (
        "LICEN", "SANITARIA", "FUNCIONAMENTO", "RESOLU",
        "ATIVIDADE", "CONCEDE", "DECLARA", "PREFEITURA",
        "MUNICIPIO", "ESTADUAL", "VIGILANCIA", "CERTIDAO",
        "BAIRRO", "LOGRADOURO", "CONSELHO", "CONTRIBUI",
        "RESPONSAVEL", "RESPONSÁVEL", "DETALHE", "SUBGRUPO",
        "AGRUPAMENTO", "OBJETO", "CNAE",
    )
    sufixo_empresa = r"(?:LTDA|EIRELI|ME|EPP|MEI|S\.?\s*A\.?|S/A)"

    def _validar(nome: str) -> str | None:
        nome = re.sub(
            r"\b(RAZ[AÃ]O\s+SOCIAL|NOME\s+FANTASIA|CONTRIBUINTE|NOME\s+EMPRESARIAL)\s*:?\s*",
            "", nome, flags=re.I,
        )
        nome = _limpar(nome)
        nome = re.sub(r"\s+", " ", nome).strip(" -.,|*:")
        if (8 < len(nome) < 160
                and sum(c.isalpha() for c in nome) > 6
                and not any(kw in nome for kw in ruins)):
            return nome
        return None

    candidatos = []

    # 2.1) Prioridade: início "RAZAO SOCIAL:" + captura GREEDY até LTDA/etc.
    padrao_rotulado = re.compile(
        r"RAZ[AÃ]O\s+SOCIAL\s*:?\s*"
        r"([A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇ&'\.][A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇ&'\.\- ]{6,180}\s+"
        + sufixo_empresa + r")\b",
        re.IGNORECASE,
    )
    for m in padrao_rotulado.finditer(up_flat):
        v = _validar(m.group(1))
        if v:
            candidatos.append(v)

    # 2.2) Genérico: qualquer sequência terminando em LTDA/etc.
    padrao_generico = re.compile(
        r"([A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇ&'\.][A-ZÁÉÍÓÚÂÊÎÔÛÃÕÇ&'\.\- ]{6,180}?\s+"
        + sufixo_empresa + r")\b"
    )
    for m in padrao_generico.finditer(up_flat):
        v = _validar(m.group(1))
        if v:
            candidatos.append(v)

    candidatos_globais.extend(candidatos)

    # 3) "RAZAO SOCIAL <texto>" — quando OCR mistura colunas, o nome real
    #    pode estar nas linhas vizinhas. Tenta juntar.
    linhas_all = texto.split("\n")
    for i, linha in enumerate(linhas_all):
        up_linha = linha.upper()
        if re.search(r"RAZ[AÃÂ]O\s+SOCIAL", up_linha):
            # Procura nas próximas linhas não-vazias por algo com letras
            for j in range(i + 1, min(len(linhas_all), i + 6)):
                cand = _limpar(linhas_all[j])
                if (5 < len(cand) < 130
                        and sum(c.isalpha() for c in cand) > 6
                        and not re.match(r"^\s*(CNPJ|NOME|LOGRADOURO|DETALHE)",
                                         cand, re.I)):
                    candidatos_globais.append(cand)
                    break

    # 4) Fallback: linha imediatamente acima/abaixo do CNPJ mascarado
    if cnpj:
        linhas = linhas_all
        cnpj_puro = re.sub(r"[^\d]", "", cnpj)
        for i, linha in enumerate(linhas):
            linha_pura = re.sub(r"[^\d]", "", linha)
            if cnpj in linha or cnpj_puro in linha_pura:
                # procura entre as 2 anteriores e 1 posterior por linha textual
                for idx in (i - 1, i - 2, i + 1):
                    if 0 <= idx < len(linhas):
                        cand = _limpar(linhas[idx])
                        # precisa ter letras, não ser só dígitos/rótulo
                        if (3 < len(cand) < 130
                                and sum(c.isalpha() for c in cand) > 4
                                and not re.match(r"^\s*CNPJ", cand, re.I)
                                and not re.match(r"^\s*\d", cand)):
                            candidatos_globais.append(cand)
                            break

    # Ranking: prefere candidatos com sufixo empresarial (LTDA/EIRELI/etc.);
    # dentro de cada grupo, o mais longo vence. Isso evita pegar títulos de
    # documento como "CERTIFICADO DE REGULARIDADE DO FGTS" por engano.
    if candidatos_globais:
        sufixo_re = re.compile(r"\b(LTDA|EIRELI|EPP|MEI|ME|S\.?\s*A\.?|S/A)\s*\.?$",
                               re.IGNORECASE)
        com_sufixo = [c for c in candidatos_globais if sufixo_re.search(c)]
        if com_sufixo:
            return max(com_sufixo, key=len)
        return max(candidatos_globais, key=len)
    return None


def _extrair_data_emissao(texto: str, tipo: str) -> str | None:
    up = texto.upper()
    # rótulos comuns (data vem APÓS o rótulo)
    padroes = [
        r"(?:EMITID[OA]\s+EM|DATA\s+DA\s+EMISS[ÃA]O|EMISS[ÃA]O)[:\s]+" + _DATA_RE,
        r"EXPEDID[OA]\s+EM\s+" + _DATA_RE,
        r"DATA\s+DE\s+EMISS[ÃA]O[:\s]+" + _DATA_RE,
        # Alvará Sanitário / documentos oficiais:
        r"DATA\s+DE\s+DEFERIMENTO[:\s]+" + _DATA_RE,
        r"DEFERIMENTO[:\s]+" + _DATA_RE,
        # "Data de expedição"
        r"DATA\s+DE\s+EXPEDI[ÇC][ÃA]O[:\s]+" + _DATA_RE,
    ]
    for p in padroes:
        m = re.search(p, up)
        if m:
            return m.group(1)

    # Rótulos-rodapé: a data aparece ANTES do rótulo (layout tabular típico
    # do Alvará Sanitário SP: "MUNICIPIO 17/03/2026 \n LOCAL DATA DE DEFERIMENTO").
    # Aumentamos a janela pra 400 chars porque no OCR real várias linhas
    # cheias de ruído separam a data do rótulo.
    padroes_antes = [
        _DATA_RE + r"[\s\S]{0,400}?DATA\s+DE\s+DEFERIMENTO",
        _DATA_RE + r"[\s\S]{0,400}?DEFERIMENTO",
    ]
    # Escolhe a data MAIS PRÓXIMA do rótulo (última ocorrência antes dele)
    for p in padroes_antes:
        matches = list(re.finditer(p, up))
        if matches:
            return matches[-1].group(1)

    # Fallback específico para Alvará Sanitário:
    # a data de emissão (deferimento) é geralmente ~1 ano ANTES da data de
    # validade. Se houver duas datas e a menor estiver coerente, usa.
    if "SANIT" in tipo.upper() or "CEVS" in up:
        datas = re.findall(_DATA_RE, texto)
        if len(datas) >= 2:
            from datetime import datetime
            parsed = []
            for d in datas:
                try:
                    parsed.append(datetime.strptime(d, "%d/%m/%Y"))
                except Exception:
                    continue
            if len(parsed) >= 2:
                parsed.sort()
                return parsed[0].strftime("%d/%m/%Y")
    return None


def _extrair_data_vencimento(texto: str, tipo: str,
                             data_emissao: str | None) -> str | None:
    up = texto.upper()
    # 1) Rótulos clássicos de validade
    padroes = [
        r"V[AÁ]LIDA?\s+AT[ÉE]\s+" + _DATA_RE,
        r"VALIDADE[:\s]+\d{1,3}\s*DIAS[^\d]*?" + _DATA_RE,  # raros
        r"VALIDADE[:\s]+" + _DATA_RE + r"\s+A\s+" + _DATA_RE,  # "Validade: dd/mm/aaaa a dd/mm/aaaa"
        r"VALIDADE[:\s]+" + _DATA_RE,
        r"VENC(?:IMENTO|E)?\s*(?:EM)?[:\s]+" + _DATA_RE,
        r"DATA\s+DE\s+VALIDADE[:\s]+" + _DATA_RE,
        r"V[AÁ]LIDO\s+AT[ÉE]\s+" + _DATA_RE,
    ]
    for p in padroes:
        m = re.search(p, up)
        if m:
            # "Validade: dd/mm a dd/mm" — o segundo é o vencimento
            if m.lastindex and m.lastindex >= 2:
                return m.group(2)
            return m.group(1)

    # 2) CNDT: 180 dias da emissão
    if "CNDT" in tipo or "TRABALHISTA" in tipo.upper():
        if data_emissao:
            from datetime import datetime, timedelta
            try:
                d = datetime.strptime(data_emissao, "%d/%m/%Y")
                return (d + timedelta(days=180)).strftime("%d/%m/%Y")
            except Exception:
                pass

    # 3) CND Federal: vale 180 dias da emissão (default)
    if "CND FEDERAL" in tipo.upper() and data_emissao:
        from datetime import datetime, timedelta
        try:
            d = datetime.strptime(data_emissao, "%d/%m/%Y")
            return (d + timedelta(days=180)).strftime("%d/%m/%Y")
        except Exception:
            pass

    # 4) Fallback genérico — pega a maior data do texto como vencimento
    datas = re.findall(_DATA_RE, texto)
    if datas:
        from datetime import datetime
        try:
            parsed = []
            for d in datas:
                try:
                    parsed.append(datetime.strptime(d, "%d/%m/%Y"))
                except Exception:
                    continue
            if parsed:
                return max(parsed).strftime("%d/%m/%Y")
        except Exception:
            pass
    return None


def _extrair_numero_documento(texto: str, tipo: str) -> str | None:
    up = texto.upper()

    # Padrões ESPECÍFICOS por tipo de documento — tentados ANTES dos genéricos.
    padroes_por_tipo: dict[str, list[str]] = {
        "Alvará Sanitário": [
            # N° CEVS — Cadastro Estadual de Vigilância Sanitária (SP).
            # Formato: 355645315-812-000003-1-0 (15 dígitos separados por hífens)
            r"N[ºOo\.°9]{0,3}\s*CEVS[:\s]+([0-9][0-9\-\.\s]{10,40})",
            r"CEVS[:\s#]+([0-9][0-9\-\.\s]{10,40})",
            r"C[ÓO]DIGO\s+CEVS[:\s]+([0-9][0-9\-\.\s]{10,40})",
        ],
        "Alvará de Funcionamento": [
            r"ALVAR[ÁA]\s+N[ºOo\.°]{1,3}[:\s]+([A-Z0-9][A-Z0-9\.\-/]{3,})",
            r"N[ºOo\.°]{1,3}\s*DO\s+ALVAR[ÁA][:\s]+([A-Z0-9\.\-/]+)",
        ],
        "Licença Ambiental": [
            r"LICEN[ÇC]A\s+N[ºOo\.°]{1,3}[:\s]+([A-Z0-9\.\-/]+)",
            r"N[ºOo\.°]{1,3}\s*DA\s+LICEN[ÇC]A[:\s]+([A-Z0-9\.\-/]+)",
        ],
    }
    for padrao in padroes_por_tipo.get(tipo, []):
        m = re.search(padrao, up)
        if m:
            num = re.sub(r"\s+", "", m.group(1)).strip(".- ")
            if len(num) >= 4 and num not in {"DA", "DO", "DE", "NULL", "NONE"}:
                return num

    # Rótulos GENÉRICOS — do mais específico pro mais genérico.
    # Cuidado: o último fallback "Nº:" só entra se os anteriores falharem,
    # pois endereços às vezes trazem "Nº 100".
    padroes = [
        r"C[ÓO]DIGO\s+DE\s+CONTROLE\s+DA\s+CERTID[ÃA]O[:\s]+([A-Z0-9\.\-/]+)",
        r"C[ÓO]DIGO\s+DE\s+AUTENTICA[ÇC][ÃA]O[:\s]+([A-Z0-9\.\-/]+)",
        r"C[ÓO]DIGO\s+DE\s+CONTROLE[:\s]+([A-Z0-9][A-Z0-9\.\-/]{3,})",
        r"CERTID[ÃA]O\s+N[ºOo\.°]{1,3}[:\s]+([A-Z0-9\.\-/]+)",
        r"N[ÚU]MERO\s+(?:DA\s+)?CERTID[ÃA]O[:\s]+([A-Z0-9\.\-/]+)",
        r"CRF[:\s#]+([A-Z0-9][A-Z0-9\.\-/]{5,})",
        r"PROTOCOLO[:\s]+([A-Z0-9\.\-/]+)",
        r"AUTENTICA[ÇC][ÃA]O[:\s]+([A-Z0-9\.\-/]+)",
        # Fallback: "Nº: XXXX" solto em qualquer lugar — só com ':' obrigatório
        r"(?:^|\n|\s)N[ºOo\.°]{1,3}\s*[:\s]\s*([A-Z0-9][A-Z0-9\.\-/]{4,})",
    ]
    for p in padroes:
        m = re.search(p, up)
        if m:
            num = m.group(1).strip(".- ")
            # Evita pegar palavras genéricas por engano
            if len(num) >= 4 and num.upper() not in {
                "DA", "DO", "DE", "CERTIDAO", "CERTIDÃO", "NULL", "NONE"
            }:
                return num
    return None


def extrair_estrutura_cnae_concla(caminho_xlsx: str) -> list[dict]:
    """
    Lê o XLSX oficial "CNAE-Subclasses X.X - Estrutura Detalhada" do IBGE
    e retorna uma lista plana de registros com todos os níveis.

    Cada item: {codigo, nivel, denominacao, secao, divisao, grupo, classe}
    Níveis: secao, divisao, grupo, classe, subclasse.

    Aceita variações do nome da aba (Subclass2.2, Subclass2.3, etc.).
    Nome esperado da aba começa com "Estrutura Det. CNAE".
    """
    import openpyxl
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True, read_only=True)

    # Encontra a aba certa
    aba_alvo = None
    for nome in wb.sheetnames:
        if "ESTRUTURA" in nome.upper() and "CNAE" in nome.upper():
            aba_alvo = nome
            break
    if not aba_alvo:
        aba_alvo = wb.sheetnames[0]  # fallback
    ws = wb[aba_alvo]

    def s(v):
        return str(v).strip() if v is not None else None

    registros: list[dict] = []
    contexto_atual = {"secao": None, "divisao": None, "grupo": None,
                      "classe": None}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Pula cabeçalhos (linhas 0-3 no formato oficial)
        if i < 4:
            continue
        row_padded = (row + (None,) * 7)[:7]
        secao, divisao, grupo, classe, subclasse, denom, _ = row_padded
        denom = str(denom).strip() if denom else ""
        if not denom:
            continue

        # Atualiza contexto hierárquico
        if s(secao):
            contexto_atual = {"secao": s(secao), "divisao": None,
                              "grupo": None, "classe": None}
            registros.append({
                "codigo": s(secao), "nivel": "secao",
                "denominacao": denom, **contexto_atual,
            })
        elif s(divisao):
            contexto_atual["divisao"] = s(divisao)
            contexto_atual["grupo"] = None
            contexto_atual["classe"] = None
            registros.append({
                "codigo": s(divisao), "nivel": "divisao",
                "denominacao": denom,
                "secao": contexto_atual["secao"],
                "divisao": s(divisao),
                "grupo": None, "classe": None,
            })
        elif s(grupo):
            contexto_atual["grupo"] = s(grupo)
            contexto_atual["classe"] = None
            registros.append({
                "codigo": s(grupo), "nivel": "grupo",
                "denominacao": denom,
                "secao": contexto_atual["secao"],
                "divisao": contexto_atual["divisao"],
                "grupo": s(grupo), "classe": None,
            })
        elif s(classe):
            contexto_atual["classe"] = s(classe)
            registros.append({
                "codigo": s(classe), "nivel": "classe",
                "denominacao": denom,
                "secao": contexto_atual["secao"],
                "divisao": contexto_atual["divisao"],
                "grupo": contexto_atual["grupo"],
                "classe": s(classe),
            })
        elif s(subclasse):
            registros.append({
                "codigo": s(subclasse), "nivel": "subclasse",
                "denominacao": denom,
                "secao": contexto_atual["secao"],
                "divisao": contexto_atual["divisao"],
                "grupo": contexto_atual["grupo"],
                "classe": contexto_atual["classe"],
            })

    wb.close()
    return registros


def extrair_dados_documento(caminho_pdf: str) -> dict:
    """
    Extrator genérico para documentos com vencimento (CND, FGTS, CNDT,
    Alvará, Licença Sanitária, etc.).

    Retorna dict com:
        tipo              string (bate com TIPOS_DOCUMENTO_VENCIMENTO)
        numero            string ou None
        cnpj              00.000.000/0000-00 ou None
        razao_social      string ou None
        data_emissao      dd/mm/aaaa ou None
        data_vencimento   dd/mm/aaaa ou None
        texto             texto bruto (debug)
        usou_ocr          bool
        idioma_ocr        'por' | 'eng' | None
    """
    texto, usou_ocr, idioma_ocr = _extrair_texto_pdf_ou_ocr(caminho_pdf)
    texto_limpo = re.sub(r"[ \t]+", " ", texto)
    up = texto_limpo.upper()

    tipo = _detectar_tipo_documento(up)
    cnpj = _extrair_cnpj(texto_limpo)
    razao = _extrair_razao_social(texto_limpo, cnpj)
    data_emi = _extrair_data_emissao(texto_limpo, tipo)
    data_venc = _extrair_data_vencimento(texto_limpo, tipo, data_emi)
    numero = _extrair_numero_documento(texto_limpo, tipo)

    return {
        "tipo": tipo,
        "numero": numero,
        "cnpj": cnpj,
        "razao_social": razao,
        "data_emissao": data_emi,
        "data_vencimento": data_venc,
        "texto": texto,
        "usou_ocr": usou_ocr,
        "idioma_ocr": idioma_ocr,
    }


def extrair_dados_auto(caminho_pdf: str) -> dict:
    """
    Despachante unificado. Rode uma vez sobre qualquer PDF:
      - Detecta o tipo (incluindo AVCB/CLCB, CND, FGTS, CNDT, Alvarás, etc.)
      - Para AVCB/CLCB, chama o extrator especializado
        (que captura ocupação IT-01, divisão, área construída...)
      - Para os demais, usa o extrator genérico.

    O dict retornado contém sempre as chaves básicas
    (tipo, numero, cnpj, razao_social, data_emissao, data_vencimento,
    usou_ocr, idioma_ocr) e, no caso AVCB/CLCB, as extras específicas
    (ocupacao, divisao, descricao_ocupacao, area_construida, altura, endereco).

    A chave adicional `destino` indica a tabela alvo:
      - "alvaras_bombeiros" → AVCB / CLCB
      - "documentos_vencimento" → todos os outros
    """
    # 1) extração genérica primeiro (barata, define tipo rapidamente)
    info_gen = extrair_dados_documento(caminho_pdf)
    tipo = info_gen.get("tipo")

    if tipo in ("AVCB", "CLCB"):
        # chama extrator especializado para pegar campos extras
        info_avcb = extrair_dados_avcb(caminho_pdf)
        # combina os dois — prioriza AVCB nos campos sobrepostos
        merged = {**info_gen, **{k: v for k, v in info_avcb.items() if v}}
        merged["destino"] = "alvaras_bombeiros"
        # garante tipo AVCB/CLCB vindo do especializado
        merged["tipo"] = info_avcb.get("tipo") or tipo
        return merged

    info_gen["destino"] = "documentos_vencimento"
    return info_gen
